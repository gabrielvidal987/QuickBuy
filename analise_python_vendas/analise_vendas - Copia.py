import mysql.connector
from mysql.connector import Error
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Conexão com o banco de dados
def criar_conexao():
    try:
        conexao = mysql.connector.connect(
            host='localhost',
            database='acompanha_pedidosschema',  # Substitua pelo nome do seu banco
            user='root',
            password='Apteste1234'
        )
        if conexao.is_connected():
            print("Conexão com MySQL bem-sucedida!")
            return conexao
    except Error as e:
        print(f"Erro ao conectar ao MySQL: {e}")
        return None

# ========== CRUD - TABELA: log_erro ==========

def ler_logs_erro(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        cursor.execute("SELECT * FROM log_erro")
    else:
        cursor.execute(f"SELECT * FROM log_erro WHERE usuario = '{usuario}' ")
    resultados = cursor.fetchall()
    res = []
    for row in resultados:
        res.append(row)
    cursor.close()
    conexao.close()
    return res

def deletar_log_erro(id_log : str | int = None, usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if id_log is None:
        cursor.execute(f"DELETE FROM log_erro WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM log_erro WHERE id_log = '{id_log}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== CRUD - TABELA: pedidos ==========

def ler_pedidos(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        cursor.execute("SELECT * FROM pedidos")
    else:
        cursor.execute(f"SELECT * FROM pedidos WHERE usuario = '{usuario}'")
    resultados = cursor.fetchall()
    res = []
    for row in resultados:
        res.append(row)
    cursor.close()
    conexao.close()
    return res

def deletar_pedido(numero_pedido : str | int = None, usuario : str = None):
    if usuario is None: return None
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if numero_pedido is None:
        cursor.execute(f"DELETE FROM pedidos WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM pedidos WHERE numero_pedido = '{numero_pedido}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== CRUD - TABELA: produtos ==========

def ler_produtos(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        # cursor.execute("SELECT * FROM produtos")
        cursor.execute("SELECT nome, qtd_vendido FROM produtos")
    else:
        cursor.execute(f"SELECT nome, qtd_vendido FROM produtos WHERE usuario = '{usuario}'")
    produtos = cursor.fetchall()
    res = []
    for p in produtos:
        res.append(p)
    cursor.close()
    conexao.close()
    return res

def deletar_produto(id_produto : str | int = None, usuario : str = None):
    if usuario is None: return None
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if id_produto is None:
        cursor.execute(f"DELETE FROM produtos WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM produtos WHERE id_produto = '{id_produto}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== CRUD - TABELA: usuarios ==========

def ler_usuarios(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        cursor.execute("SELECT * FROM usuarios")
    else:
        cursor.execute(f"SELECT * FROM usuarios WHERE usuario = '{usuario}'")
    usuarios = cursor.fetchall()
    res = []
    for user in usuarios:
        res.append(user)
    cursor.close()
    conexao.close()
    return res

def deletar_usuario(id_sys : str | int = None, usuario : str = None):
    if usuario is None: return None
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if id_sys is None:
        cursor.execute(f"DELETE FROM usuarios WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM usuarios WHERE id_sys = '{id_sys}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== GERAR PLANILHA ============
# Função para transformar booleanos (0/1) em "Sim"/"Não"
def booleano_para_sim_nao(valor):
    return "Sim" if valor == 1 else "Não"

# Função principal que gera o relatório
def gerar_relatorio_excel():
    # ===================== tira os relatorios ===================
    pedidos = ler_pedidos()
    produtos = ler_produtos()
    # ===================== Aba 1 - Vendas =====================

    # Mapeamento de colunas para nomes amigáveis
    colunas_pedidos = {
        "numero_pedido": "Nº do Pedido",
        "nome_cliente": "Nome do Cliente",
        "endereco": "Endereço",
        "produtos_nome": "Produtos",
        "observacoes": "Observações",
        "hora_pedido": "Hora do Pedido",
        "hora_ficou_pronto": "Hora Ficou Pronto",
        "valor_total": "Valor Total (R$)",
        "forma_pag": "Forma de Pagamento",
        "pagamento_aprovado": "Pagamento Aprovado",
        "pedido_pronto": "Pedido Pronto",
        "usuario": "Usuário",
        "delivery": "Delivery"
    }

    
    # ===================== Aba 2 - Produtos =====================

    # Criar planilha
    wb = Workbook()

    # -------- Aba 1: Vendas --------
    ws_vendas = wb.active
    ws_vendas.title = "Vendas"

    headers_vendas = [colunas_pedidos[col] for col in colunas_pedidos]
    ws_vendas.append(headers_vendas)

    for pedido in pedidos:
        linha = []
        for col in colunas_pedidos:
            valor = pedido[col]
            if col in ["pagamento_aprovado", "pedido_pronto", "delivery"]:
                valor = booleano_para_sim_nao(valor)
            linha.append(valor)
        ws_vendas.append(linha)

    # Auto ajustar colunas
    for col in ws_vendas.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_vendas.column_dimensions[col_letter].width = max_len + 2

    # -------- Aba 2: Produtos Vendidos --------
    ws_produtos = wb.create_sheet("Produtos Vendidos")
    ws_produtos.append(["Produto", "Quantidade Vendida"])

    for produto in produtos:
        ws_produtos.append([produto["nome"], produto["qtd_vendido"]])

    for col in ws_produtos.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_produtos.column_dimensions[col_letter].width = max_len + 2

    # Salvar o arquivo
    nome_arquivo = "relatorio_vendas.xlsx"
    wb.save(nome_arquivo)
    print(f"Relatório gerado com sucesso: {nome_arquivo}")


# =======================
# EXEMPLO DE USO
# =======================
if __name__ == "__main__":
    # logs = ler_logs_erro()
    gerar_relatorio_excel()
    pass
