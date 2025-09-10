import mysql.connector
from mysql.connector import Error
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from collections import Counter

# Conexão com o banco de dados
def criar_conexao():
    try:
        conexao = mysql.connector.connect(
            host='localhost',
            database='acompanha_pedidosschema',  # Substitua pelo nome do seu banco
            user='root',
            password='Apteste1234'
        )
        if conexao.is_connected():
            return conexao
    except Error as e:
        print(f"Erro ao conectar ao MySQL: {e}")
        return None

# ========== CRUD - TABELA: log_erro ==========

def ler_logs_erro(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        cursor.execute("SELECT * FROM log_erro")
    else:
        cursor.execute(f"SELECT * FROM log_erro WHERE usuario = '{usuario}' ")
    resultados = cursor.fetchall()
    res = []
    for row in resultados:
        res.append(row)
    cursor.close()
    conexao.close()
    return res

def deletar_log_erro(id_log : str | int = None, usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if id_log is None:
        cursor.execute(f"DELETE FROM log_erro WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM log_erro WHERE id_log = '{id_log}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== CRUD - TABELA: pedidos ==========

def ler_pedidos(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        cursor.execute("SELECT * FROM pedidos")
    else:
        cursor.execute(f"SELECT * FROM pedidos WHERE usuario = '{usuario}'")
    resultados = cursor.fetchall()
    res = []
    for row in resultados:
        res.append(row)
    cursor.close()
    conexao.close()
    return res

def deletar_pedido(numero_pedido : str | int = None, usuario : str = None):
    if usuario is None: return None
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if numero_pedido is None:
        cursor.execute(f"DELETE FROM pedidos WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM pedidos WHERE numero_pedido = '{numero_pedido}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== CRUD - TABELA: produtos ==========

def ler_produtos(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        # cursor.execute("SELECT * FROM produtos")
        cursor.execute("SELECT nome, qtd_vendido FROM produtos")
    else:
        cursor.execute(f"SELECT nome, qtd_vendido FROM produtos WHERE usuario = '{usuario}'")
    produtos = cursor.fetchall()
    res = []
    for p in produtos:
        res.append(p)
    cursor.close()
    conexao.close()
    return res

def deletar_produto(id_produto : str | int = None, usuario : str = None):
    if usuario is None: return None
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if id_produto is None:
        cursor.execute(f"DELETE FROM produtos WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM produtos WHERE id_produto = '{id_produto}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== CRUD - TABELA: usuarios ==========

def ler_usuarios(usuario : str = None):
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor(dictionary=True)
    if usuario is None:
        cursor.execute("SELECT * FROM usuarios")
    else:
        cursor.execute(f"SELECT * FROM usuarios WHERE usuario = '{usuario}'")
    usuarios = cursor.fetchall()
    res = []
    for user in usuarios:
        res.append(user)
    cursor.close()
    conexao.close()
    return res

def deletar_usuario(id_sys : str | int = None, usuario : str = None):
    if usuario is None: return None
    conexao = criar_conexao()
    if not conexao: return None
    cursor = conexao.cursor()
    if id_sys is None:
        cursor.execute(f"DELETE FROM usuarios WHERE usuario = '{usuario}' ")
    else:
        cursor.execute(f"DELETE FROM usuarios WHERE id_sys = '{id_sys}' AND usuario = '{usuario}' ")
    conexao.commit()
    cursor.close()
    conexao.close()
    return True

# ========== GERAR PLANILHA ============
# Função para transformar booleanos (0/1) em "Sim"/"Não"
def booleano_para_sim_nao(valor):
    return "Sim" if valor == 1 else "Não"

# Função principal que gera o relatório
def gerar_relatorio_excel(usuario : str = None):
    # ===================== tira os relatorios ===================
    pedidos = ler_pedidos(usuario)
    produtos = ler_produtos(usuario)

    # Contar pedidos por cliente
    contagem_clientes = Counter()
    for p in pedidos:
        nome = p["nome_cliente"]
        contagem_clientes[nome] += 1

    top_clientes = contagem_clientes.most_common(10)

    # Top produtos mais vendidos (já vem prontos da tabela)
    top_produtos = sorted(produtos, key=lambda x: x["qtd_vendido"], reverse=True)

    # ===================== Criar Planilha =====================
    wb = Workbook()

    # -------- Aba 1: Vendas --------
    ws_vendas = wb.active
    ws_vendas.title = "Vendas"

    colunas_pedidos = {
        "numero_pedido": "Nº do Pedido",
        "nome_cliente": "Nome do Cliente",
        "endereco": "Endereço",
        "produtos_nome": "Produtos",
        "observacoes": "Observações",
        "hora_pedido": "Hora do Pedido",
        "hora_ficou_pronto": "Hora Ficou Pronto",
        "valor_total": "Valor Total (R$)",
        "forma_pag": "Forma de Pagamento",
        "pagamento_aprovado": "Pagamento Aprovado",
        "pedido_pronto": "Pedido Pronto",
        "usuario": "Usuário",
        "delivery": "Delivery"
    }

    headers_vendas = [colunas_pedidos[col] for col in colunas_pedidos]
    ws_vendas.append(headers_vendas)

    for pedido in pedidos:
        linha = []
        for col in colunas_pedidos:
            valor = pedido[col]
            if col in ["pagamento_aprovado", "pedido_pronto", "delivery"]:
                valor = booleano_para_sim_nao(valor)
            linha.append(valor)
        ws_vendas.append(linha)

    for col in ws_vendas.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_vendas.column_dimensions[col_letter].width = max_len + 2

    # -------- Aba 2: Produtos Vendidos --------
    ws_prod = wb.create_sheet("Produtos Vendidos")
    ws_prod.append(["Produto", "Quantidade Vendida"])

    for produto in produtos:
        ws_prod.append([produto["nome"], produto["qtd_vendido"]])

    for col in ws_prod.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_prod.column_dimensions[col_letter].width = max_len + 2

    # -------- Aba 3: Top Clientes (com gráfico) --------
    ws_top_clientes = wb.create_sheet("Top Clientes")
    ws_top_clientes.append(["Cliente", "Nº de Pedidos"])

    for cliente, qtd in top_clientes:
        ws_top_clientes.append([cliente, qtd])

    # Gráfico de Pizza para Top Clientes
    chart1 = PieChart()
    chart1.title = "Top 10 Clientes com Mais Pedidos"
    data = Reference(ws_top_clientes, min_col=2, min_row=1, max_row=len(top_clientes) + 1)
    labels = Reference(ws_top_clientes, min_col=1, min_row=2, max_row=len(top_clientes) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(labels)
    chart1.height = 10
    chart1.width = 10
    ws_top_clientes.add_chart(chart1, "D2")

    # -------- Aba 4: Top Produtos (com gráfico) --------
    ws_top_produtos = wb.create_sheet("Top Produtos")
    ws_top_produtos.append(["Produto", "Quantidade Vendida"])

    for produto in top_produtos[:10]:
        ws_top_produtos.append([produto["nome"], produto["qtd_vendido"]])

    # Gráfico de Pizza para Produtos mais vendidos
    chart2 = PieChart()
    chart2.title = "Top 10 Produtos Mais Vendidos"
    data = Reference(ws_top_produtos, min_col=2, min_row=1, max_row=11)
    labels = Reference(ws_top_produtos, min_col=1, min_row=2, max_row=11)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(labels)
    chart2.height = 10
    chart2.width = 10
    ws_top_produtos.add_chart(chart2, "D2")

    # -------- Salvar Planilha --------
    if not usuario: usuario = 'Anonimo'
    nome_arquivo = f"relatorio_vendas_{usuario}.xlsx"
    wb.save(nome_arquivo)
    print(f"Relatório com gráficos gerado: {nome_arquivo}")

# =======================
# EXEMPLO DE USO
# =======================
if __name__ == "__main__":
    # logs = ler_logs_erro()
    print('='*30)
    print('     PROGRAMA PARA REALIZAR ANALISE SOBRE VENDAS')
    print('     ANALISANDO...')
    print('='*30)
    user = input('\n\nDigite o nome do usuario conforme a senha:\nR->')
    gerar_relatorio_excel(user)
    pass
